from __future__ import annotations

from dataclasses import dataclass
from typing import List, Tuple

import openpyxl
import pandas as pd


@dataclass(frozen=True)
class CorrInputs:
    """Minimal correlation inputs extracted from the Excel 'correlaciones' sheet."""
    theta: float
    scenario_scale: float


def load_corr_inputs_from_excel(xlsx_path: str, sheet_name: str = "correlaciones") -> CorrInputs:
    """Load theta and scenario_scale from the provided Excel file.

    Excel locations (as in Delta_GIRR_Sol.xlsx):
        - theta         : correlaciones!B2
        - scenario_scale: correlaciones!Y2
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    theta = float(ws["B2"].value)
    scenario = float(ws["Y2"].value)
    return CorrInputs(theta=theta, scenario_scale=scenario)


def parse_delta_sheet_from_excel(
    xlsx_path: str,
    sheet_name: str,
    default_ccy: str = "EUR",
) -> pd.DataFrame:
    """Parse one of the example sheets into a normalized sensitivity table.

    It expects the same layout as Delta_GIRR_Sol.xlsx example tabs.

    Output columns:
        ccy, curve, tenor, ftype, delta
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # Find tenor header row: cell C == 'Tenor'
    tenor_row = None
    for r in range(1, 40):
        v = ws.cell(r, 3).value
        if isinstance(v, str) and v.strip().lower() == "tenor":
            tenor_row = r
            break
    if tenor_row is None:
        raise ValueError(f"Tenor header row not found in sheet={sheet_name}")

    # Tenors: row tenor_row, columns D.. until blank
    tenors: List[float] = []
    tenor_cols: List[int] = []
    for c in range(4, ws.max_column + 1):
        v = ws.cell(tenor_row, c).value
        if v is None or v == "":
            break
        tenors.append(float(v))
        tenor_cols.append(c)

    # Delta rows: below tenor row, until first empty line after we started collecting.
    rows = []
    started = False
    for r in range(tenor_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        curve = ws.cell(r, 2).value

        if label is None or str(label).strip() == "":
            if started:
                break
            continue

        label_s = str(label).strip()
        is_delta_row = label_s.lower().startswith("curva") or ("inflacion" in label_s.lower())
        if not is_delta_row:
            continue

        started = True

        # Determine currency
        parts = label_s.split()
        ccy = None
        for token in reversed(parts):
            if token.isalpha() and len(token) == 3:
                ccy = token.upper()
                break
        if ccy is None:
            ccy = default_ccy

        ftype = "inflation" if "inflacion" in label_s.lower() else "rate"
        curve_s = str(curve).strip() if curve is not None else ("Inflacion" if ftype == "inflation" else "UNKNOWN")

        for tenor, c in zip(tenors, tenor_cols):
            delta = ws.cell(r, c).value
            rows.append(
                {
                    "ccy": ccy,
                    "curve": curve_s,
                    "tenor": float(tenor),
                    "ftype": ftype,
                    "delta": float(delta) if delta is not None else 0.0,
                }
            )

    return pd.DataFrame(rows)
